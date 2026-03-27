import pandas as pd
from openpyxl import load_workbook
import os
import json
from datetime import datetime
import shutil


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

TEMPLATE_PATH = os.path.join(BASE_DIR, "templates", "master_template.xlsx")
CONFIG_PATH = os.path.join(BASE_DIR, "configs", "mapping.json")
INBOX_DIR = os.path.join(BASE_DIR, "inbox")
OUTBOX_DIR = os.path.join(BASE_DIR, "outbox")
ARCHIVE_DIR = os.path.join(BASE_DIR, "processed")

def process_data():
    try:
       
        if not os.path.exists(CONFIG_PATH):
            print("mapping.json not found")
            return None

        with open(CONFIG_PATH, 'r') as f:
            cfg = json.load(f)

      
        folders = [
            f for f in os.listdir(INBOX_DIR)
            if os.path.isdir(os.path.join(INBOX_DIR, f))
        ]

        if not folders:
            print("No folders in inbox")
            return None

        latest_folder = max(
            folders,
            key=lambda f: os.path.getctime(os.path.join(INBOX_DIR, f))
        )

        folder_path = os.path.join(INBOX_DIR, latest_folder)

       
        csv_files = [f for f in os.listdir(folder_path) if f.endswith('.csv')]
        if not csv_files:
            print("No CSV in folder")
            return None

        file_path = os.path.join(folder_path, csv_files[0])
        df_raw = pd.read_csv(file_path)

  
        now = datetime.now()
        current_month_num = now.month
        current_month_name = now.strftime("%B")

        processed_rows = []

        for _, row in df_raw.iterrows():
            is_income = row['Income'] > 0
            amount = row['Income'] if is_income else row['Expense']
            txn_type = "Income" if is_income else "Expenses"

            raw_cat = str(row['Category']).strip()
            clean_cat = cfg['synonyms'].get(raw_cat.lower(), raw_cat.capitalize())

            
            try:
                parts = str(row['Date']).split('/')
                day = int(parts[0])
                year = int(parts[2])

                if year < 100:
                    year += 2000

                fixed_date = datetime(year, current_month_num, day)

            except Exception:
                print(f"Skipping invalid date: {row['Date']}")
                continue

            processed_rows.append({
                'Date': fixed_date,
                'Type': txn_type,
                'Category': clean_cat,
                'Amount': amount,
                'Description': row['Notes'] if pd.notna(row['Notes']) else ""
            })

        
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

       
        ws['F5'] = current_month_name

        start_row = 9

        for i, item in enumerate(processed_rows):
            r = start_row + i
            ws.cell(row=r, column=2).value = item['Date']
            ws.cell(row=r, column=3).value = item['Type']
            ws.cell(row=r, column=4).value = item['Category']
            ws.cell(row=r, column=5).value = item['Amount']
            ws.cell(row=r, column=6).value = item['Description']

      
        if not os.path.exists(OUTBOX_DIR):
            os.makedirs(OUTBOX_DIR)

        
        output_file = f"Expenses_{current_month_name}.xlsx"
        output_path = os.path.join(OUTBOX_DIR, output_file)

        wb.save(output_path)

        
        if not os.path.exists(ARCHIVE_DIR):
            os.makedirs(ARCHIVE_DIR)

        shutil.move(
            file_path,
            os.path.join(ARCHIVE_DIR, os.path.basename(file_path))
        )

        print(f"Processed: {latest_folder}")
        return output_path

    except Exception as e:
        print(f"Error: {e}")
        return None
